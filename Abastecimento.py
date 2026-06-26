import os
import time
import calendar
import requests
import smtplib
import psycopg2
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed
from tqdm import tqdm
from email.message import EmailMessage
from dotenv import load_dotenv
from openpyxl import load_workbook 

# Carrega as senhas (no GitHub ele ignora isso e pega dos Secrets direto)
load_dotenv()

# =====================================================================
# 1. CONFIGURAÇÕES E CREDENCIAIS
# =====================================================================
# Gobrax
TOKEN_GOBRAX = os.getenv("TOKEN_GOBRAX")
URL_GOBRAX = "https://gateway-v3.gobrax.com.br:8889/api/v1/vehicle-statistics"
DB_HOST = os.getenv("DB_HOST")
DB_PORT = os.getenv("DB_PORT")
DB_NAME = os.getenv("DB_NAME")
DB_USER = os.getenv("DB_USER")
DB_PASSWORD = os.getenv("DB_PASSWORD")

# Ticketlog
AUTHORIZATION_TICKET = os.getenv("AUTHORIZATION_TICKET")
CODIGO_CLIENTE = os.getenv("CODIGO_CLIENTE")
URL_TICKET = "https://srv1.ticketlog.com.br/ticketlog-servicos/ebs/transacaoVeiculo/search"

# Email (Para o Power Automate capturar)
EMAIL_REMETENTE = os.getenv("EMAIL_REMETENTE")
SENHA_EMAIL = os.getenv("SENHA_EMAIL") 
EMAIL_DESTINO = os.getenv("EMAIL_DESTINO") 

MAX_WORKERS = 5

# =====================================================================
# 2. FUNÇÕES DE EXTRAÇÃO
# =====================================================================
def extrair_gobrax(data_inicio_str, data_fim_str):
    print("\n[1/3] 📡 Extraindo telemetria (Gobrax)...")
    
    conn = psycopg2.connect(host=DB_HOST, port=DB_PORT, database=DB_NAME, user=DB_USER, password=DB_PASSWORD)
    sql_placas = """
    SELECT DISTINCT ON (v."PLACA") v."PLACA", va2."DESCRICAO" AS agrupamento
    FROM veiculo.veiculo v
    JOIN veiculo.veiculo_tipo_carroceria vtc ON vtc."TIPO_CARROCERIA_ID" = vtc."TIPO_CARROCERIA_ID"
    JOIN veiculo.veiculo_agrupamento va2 ON va2."AGRUPAMENTO_ID" = vtc."AGRUPAMENTO_ID"
    JOIN veiculo.veiculo_modalidade_atual vma ON v."PLACA" = vma."PLACA"
    WHERE vma."MODALIDADE" = 'FROTA' AND v."PLACA" IS NOT NULL
      AND va2."DESCRICAO" NOT LIKE '%SEMI REBOQUE%' AND va2."DESCRICAO" NOT LIKE '%TERCEIRO%'
    ORDER BY v."PLACA", v."VEICULO_ID" DESC
    """
    df_placas = pd.read_sql_query(sql_placas, conn)
    conn.close()

    def consultar_placa(placa, agrupamento):
        headers = {"Authorization": f"Bearer {TOKEN_GOBRAX}", "Accept": "application/json"}
        params = {"startDate": data_inicio_str, "endDate": data_fim_str, "vehicleIdentification": placa}
        
        for _ in range(3):
            try:
                r = requests.get(URL_GOBRAX, headers=headers, params=params, timeout=30)
                if r.status_code == 200:
                    records = r.json().get("records", [])
                    if records:
                        item = records[0]
                        return {
                            "PLACA": placa, "AGRUPAMENTO": agrupamento,
                            "KM_RODADO": item.get("totalMileage", 0),
                            "CONSUMO_TOTAL": item.get("totalConsumption", 0)
                        }
                    break
            except:
                time.sleep(1)
        return {"PLACA": placa, "AGRUPAMENTO": agrupamento, "KM_RODADO": 0, "CONSUMO_TOTAL": 0}

    resultado = []
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = [executor.submit(consultar_placa, row.PLACA, row.agrupamento) for row in df_placas.itertuples()]
        for f in tqdm(as_completed(futures), total=len(futures), desc="Gobrax"):
            resultado.append(f.result())
            
    df = pd.DataFrame(resultado)
    df["PLACA"] = df["PLACA"].str.upper().str.strip()
    return df

def extrair_ticketlog(data_inicio, data_fim):
    print("\n[2/3] 💳 Extraindo abastecimentos completos (Ticketlog)...")
    
    dados = {}
    data_atual = data_inicio
    
    while data_atual <= data_fim:
        inicio_dia = data_atual.strftime("%Y-%m-%dT00:00:00")
        fim_dia = data_atual.strftime("%Y-%m-%dT23:59:59")
        
        for considerar in ["V", "T"]:
            payload = {
                "codigoCliente": CODIGO_CLIENTE, "codigoTipoCartao": 4,
                "dataTransacaoInicial": inicio_dia, "dataTransacaoFinal": fim_dia,
                "considerarTransacao": considerar, "ordem": "S", "validacao": "S"
            }
            headers = {"Content-Type": "application/json", "Authorization": AUTHORIZATION_TICKET}
            
            try:
                r = requests.post(URL_TICKET, json=payload, headers=headers, timeout=30)
                if r.status_code == 200 and r.json().get("sucesso"):
                    for t in r.json().get("transacoes", []):
                        key = f"{t.get('codigoTransacao')}_{t.get('considerarTransacao')}"
                        if key not in dados:
                            t["considerarTransacao"] = considerar
                            dados[key] = t
            except:
                pass
        data_atual += timedelta(days=1)
        
    df = pd.DataFrame(list(dados.values()))
    if df.empty: return df
    
    df["dataTransacao"] = pd.to_datetime(df["dataTransacao"])
    df = df.sort_values(by=["placa", "dataTransacao"])
    df["KM Rodado"] = df.groupby("placa")["quilometragem"].diff().fillna(0)
    
    df["Data"] = df["dataTransacao"].dt.strftime("%Y-%m-%d")
    df["Hora"] = df["dataTransacao"].dt.strftime("%H:%M:%S")
    df["placa"] = df["placa"].str.upper().str.strip()
    
    colunas_completas = [
        "Data", "Hora", "placa", "quilometragem", "KM Rodado",
        "nomeReduzidoEstabelecimento", "nomeCidade", "uf",
        "tipoCombustivel", "litros", "valorTransacao",
        "valorLitro", "numeroCartao", "considerarTransacao"
    ]
    
    colunas_presentes = [c for c in colunas_completas if c in df.columns]
    
    return df[colunas_presentes].rename(columns={"placa": "PLACA"})

# =====================================================================
# 3. REGRAS DE NEGÓCIO E CONSOLIDAÇÃO
# =====================================================================
def categorizar_combustivel(tipo):
    tipo = str(tipo).upper()
    if "ARLA" in tipo: return "ARLA 32"
    elif "GNV" in tipo: return "GNV"
    else: return "DIESEL"

def processar_relatorio(ano, mes):
    primeiro_dia = datetime(ano, mes, 1)
    ultimo_dia = datetime(ano, mes, calendar.monthrange(ano, mes)[1], 23, 59, 59)
    
    df_gobrax = extrair_gobrax(primeiro_dia.strftime("%Y-%m-%d %H:%M:%S"), ultimo_dia.strftime("%Y-%m-%d %H:%M:%S"))
    df_ticket_completo = extrair_ticketlog(primeiro_dia, ultimo_dia)
    
    if df_ticket_completo.empty and df_gobrax.empty:
        print("⚠️ Sem dados para processar.")
        return None
        
    print("\n[3/3] 🧠 Processando inteligência de dados e regras de negócio...")
    
    df_ticket = df_ticket_completo.copy()
    
    # Tratamento numérico para evitar erros na Pivot Table
    df_ticket["litros"] = pd.to_numeric(df_ticket["litros"], errors='coerce').fillna(0)
    df_ticket["valorTransacao"] = pd.to_numeric(df_ticket["valorTransacao"], errors='coerce').fillna(0)
    
    df_ticket["Categoria"] = df_ticket["tipoCombustivel"].apply(categorizar_combustivel)
    
    df_ticket_sem_arla = df_ticket[df_ticket["Categoria"] != "ARLA 32"]
    df_ticket_geral = df_ticket_sem_arla.groupby("PLACA").agg(Km_Ticket=("KM Rodado", "sum"), Consumo_Ticket=("litros", "sum")).reset_index()
    df_ticket_fin = df_ticket.groupby("PLACA").agg(Custo_Total_Ticket=("valorTransacao", "sum")).reset_index()
    df_ticket_geral = pd.merge(df_ticket_geral, df_ticket_fin, on="PLACA", how="outer")
    
    # Pivot de Combustíveis (Separa GNV, Diesel e Arla)
    df_comb = df_ticket.pivot_table(index="PLACA", columns="Categoria", values=["litros", "valorTransacao"], aggfunc="sum", fill_value=0)
    df_comb.columns = [f"{col[1]} (R$)" if col[0] == "valorTransacao" else f"{col[1]} (L)" for col in df_comb.columns]
    df_comb = df_comb.reset_index().drop(columns=["ARLA 32 (L)"], errors='ignore')

    df_final = pd.merge(df_gobrax, df_ticket_geral, on="PLACA", how="outer").fillna(0)
    df_final = pd.merge(df_final, df_comb, on="PLACA", how="left").fillna(0)
    
    colunas_gnv = [col for col in df_final.columns if 'GNV' in col.upper() and '(L)' in col]
    df_final['Tem_GNV'] = df_final[colunas_gnv].sum(axis=1) > 0 if colunas_gnv else False

    df_final["KM Rodado Real"] = np.where(df_final["Tem_GNV"], df_final["Km_Ticket"],
                                 np.where(df_final["KM_RODADO"] > 0, df_final["KM_RODADO"], df_final["Km_Ticket"]))
    
    df_final["Consumo Real (L)"] = np.where(df_final["Tem_GNV"], df_final["Consumo_Ticket"],
                                   np.where(df_final["CONSUMO_TOTAL"] > 0, df_final["CONSUMO_TOTAL"], df_final["Consumo_Ticket"]))

    df_final["Média KM/L"] = np.where(df_final["Consumo Real (L)"] > 0, df_final["KM Rodado Real"] / df_final["Consumo Real (L)"], 0)
    df_final["Preço Médio Litro (R$)"] = np.where(df_final["Consumo_Ticket"] > 0, df_final["Custo_Total_Ticket"] / df_final["Consumo_Ticket"], 0)
    
    df_final["Fonte do Dado"] = np.where(df_final["Tem_GNV"], "Ticketlog (Regra GNV)",
                                np.where(df_final["CONSUMO_TOTAL"] > 0, "Gobrax", "Ticketlog (Faltou Gobrax)"))

    df_final["AGRUPAMENTO"] = df_final["AGRUPAMENTO"].replace(0, "NÃO INFORMADO")
    df_final.insert(0, "ANO", ano)
    df_final.insert(1, "MES", mes)
    
    cols_principais = ["ANO", "MES", "PLACA", "AGRUPAMENTO", "KM Rodado Real", "Consumo Real (L)", "Média KM/L", "Preço Médio Litro (R$)", "Fonte do Dado"]
    
    for c in ["GNV (L)", "GNV (R$)", "DIESEL (L)", "DIESEL (R$)", "ARLA 32 (R$)"]:
        if c not in df_final.columns:
            df_final[c] = 0.0

    lista_combustiveis = ["DIESEL (L)", "DIESEL (R$)", "GNV (L)", "GNV (R$)", "ARLA 32 (R$)"]
    cols_finais = cols_principais + lista_combustiveis
    
    df_relatorio = df_final[cols_finais].sort_values(by="PLACA").round(2)
    
    # =====================================================================
    # 3.1 SALVANDO E APLICANDO FORMATAÇÃO DE MOEDA/DECIMAIS BRASILEIRA
    # =====================================================================
    nome_arquivo = f"Relatorio_Fechamento_{ano}_{mes:02d}.xlsx"
    
    # Etapa A: Salvar os dados puros via Pandas
    with pd.ExcelWriter(nome_arquivo, engine="openpyxl") as writer:
        df_ticket_completo.sort_values(by="Data").to_excel(writer, sheet_name="Aba 1 - Ticketlog Bruto", index=False)
        df_relatorio.to_excel(writer, sheet_name="Aba 2 - Relatorio Real", index=False)
        
    # Etapa B: Arrumar o visual do Excel (pontos, vírgulas e o 0E-2)
    wb = load_workbook(nome_arquivo)
    
    # TRUQUE INFALÍVEL: Pegamos a lista exata de nomes que o próprio sistema gerou
    nomes_abas = wb.sheetnames
    ws_bruto = wb[nomes_abas[0]]
    ws_real = wb[nomes_abas[1]]
    
    # Aba 2 - Consolidada
    for row in ws_real.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                if cell.value == 0:
                    cell.value = 0.0  # Limpa o bug do 0E-2
                cell.number_format = '#,##0.00'  # Força 2 casas decimais com separador
                
    # Aba 1 - Bruta
    for row in ws_bruto.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                if cell.value == 0:
                    cell.value = 0.0
                cell.number_format = '#,##0.00'

    wb.save(nome_arquivo)
    return nome_arquivo

# =====================================================================
# 4. ENVIO DE E-MAIL (Gatilho para o Power Automate)
# =====================================================================
def enviar_email(caminho_arquivo):
    print(f"\n📧 Enviando {caminho_arquivo} por e-mail...")
    try:
        msg = EmailMessage()
        msg['Subject'] = f"Relatório Fechamento - {caminho_arquivo}"
        msg['From'] = EMAIL_REMETENTE
        msg['To'] = EMAIL_DESTINO
        msg.set_content("Segue em anexo o relatório automatizado consolidado.\n\nGerado via GitHub Actions.")

        with open(caminho_arquivo, 'rb') as f:
            file_data = f.read()
        msg.add_attachment(file_data, maintype='application', subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename=caminho_arquivo)

        # Usando Gmail. Se for Office 365, basta trocar para 'smtp.office365.com' e porta 587 (com STARTTLS).
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
            smtp.login(EMAIL_REMETENTE, SENHA_EMAIL)
            smtp.send_message(msg)
        
        print("✅ E-mail enviado com sucesso! O Power Automate já pode atuar.")
    except Exception as e:
        print(f"⚠️ Erro ao enviar o e-mail: {e}")

# =====================================================================
# 5. EXECUÇÃO PRINCIPAL
# =====================================================================
if __name__ == "__main__":
    data_referencia = datetime.now() - timedelta(days=5)
    ano_alvo = data_referencia.year
    mes_alvo = data_referencia.month
    
    print(f"🚀 INICIANDO ROBÔ DE FECHAMENTO | Ref: {ano_alvo}-{mes_alvo:02d}")
    
    arquivo_gerado = processar_relatorio(ano_alvo, mes_alvo)
    
    if arquivo_gerado:
        enviar_email(arquivo_gerado)
        print("🎉 Processo 100% finalizado com sucesso!")
